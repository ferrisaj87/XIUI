#!/usr/bin/env python3
"""
Sync BST jug spreadsheet Ready column against retail wiki data (MediaWiki API).

Primary: FFXIclopedia (ffxiclopedia.fandom.com).
Fallback when a page does not exist: BG-Wiki (`bg-wiki.com`) — still retail-era FFXI wiki data.

Parses beastmaster Jug wikitables (`Ready` column last `<td><ul>...</ul>`), respecting
`/bstpet` indices in `<b>Name (index)</b>` where present, charge tier usually from trailing
“(N Charges)”.
"""

from __future__ import annotations

import json
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from bs4 import BeautifulSoup

APIS: list[tuple[str, str]] = [
    ("ffxiclopedia", "https://ffxiclopedia.fandom.com/api.php"),
    ("bg-wiki", "https://www.bg-wiki.com/api.php"),
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
    ),
    "Accept-Encoding": "identity",
}


def _fetch_json(base: str, query: dict) -> dict:
    qs = urllib.parse.urlencode(query)
    url = f"{base}?{qs}"
    last_err = None
    for attempt in range(5):
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            return json.loads(urllib.request.urlopen(req, timeout=60).read())
        except urllib.error.HTTPError as e:
            last_err = e
            time.sleep(min(14.0, 2.8 * (attempt + 1)))
        except OSError as e:
            last_err = e
            time.sleep(2.8 * (attempt + 1))
    raise RuntimeError(last_err)


def mw_parse_fragment(page_title: str) -> tuple[str | None, str | None]:
    """Return (html_fragment, wiki_label) using first wiki that parses successfully."""
    for label, api in APIS:
        j = _fetch_json(api, {"action": "parse", "format": "json", "prop": "text", "page": page_title})
        p = j.get("parse")
        if not p:
            continue
        html = p["text"].get("*")
        if isinstance(html, str) and html.strip():
            redirect_only = "<div class=\"redirectMsg\"" in html and "Redirect to:" in html
            if redirect_only and api.endswith("bg-wiki.com/api.php"):
                continue  # Droopy Dortwin redirects to Jug item on BG-Wiki — try next / Fandom wins
            return html, label
    return None, None


def canonical_exists(title: str, api: str) -> str | None:
    j = _fetch_json(api, {"action": "query", "format": "json", "titles": title})
    pages = j.get("query", {}).get("pages", {})
    for _pid, pdata in pages.items():
        if pdata.get("missing") is True:
            return None
        return pdata.get("title")
    return None


def resolve_wiki_title(excel_name: str) -> str | None:
    cand = excel_name.strip()
    for api in (_a[1] for _a in APIS):
        t = canonical_exists(cand, api)
        if t:
            return t
        t = canonical_exists(cand.replace(" ", "_"), api)
        if t:
            return t
        if "-" in cand:
            t = canonical_exists(cand.replace("-", ""), api)
            if t:
                return t
        t = canonical_exists(cand.replace("-", " "), api)
        if t:
            return t
    return fuzzy_title_search_ffxiclopedia_only(cand.split("(")[0].strip())


def fuzzy_title_search_ffxiclopedia_only(q: str) -> str | None:
    api = APIS[0][1]
    j = _fetch_json(
        api,
        {
            "action": "query",
            "format": "json",
            "list": "search",
            "srnamespace": "0",
            "srlimit": "12",
            "srsearch": q,
        },
    )
    hits = [h["title"] for h in j.get("query", {}).get("search", [])]
    word0 = q.strip().lower().split()[0] if q.strip() else ""
    for order in ("prefix", "any"):
        for h in hits:
            hl = h.lower()
            ok = hl.startswith(word0) if order == "prefix" else word0[:4] in hl
            if not ok:
                continue
            ct = canonical_exists(h, APIS[0][1])
            if ct:
                return ct
            ct = canonical_exists(h, APIS[1][1])
            if ct:
                return ct
    return None


def core_name(txt: str) -> str:
    return re.sub(r"\s*\([^)]*\)\s*$", "", txt).strip().lower()


def alnum_key(txt: str) -> str:
    return re.sub(r"[^a-z0-9]", "", txt.lower())


def row_matches_familiar(first_td_text: str, familiar: str) -> bool:
    """Match Wiki first column (allows RhymingShizuna vs Rhyming Shizuna)."""
    a = alnum_key(core_name(familiar))
    b = alnum_key(core_name(first_td_text))
    if len(a) < 6 or len(b) < 6:
        return False
    return a == b


def extract_charge(text: str) -> int | None:
    tl = text.replace("&nbsp;", " ")
    # Prefer Requires N charge(s)
    mreq = re.search(r"requires\s+(\d+)\s+charges?", tl, re.I)
    if mreq:
        return int(mreq.group(1))
    # Trailing parentheses on line (tier): take LAST (N Charges) occurrence
    chs = list(re.finditer(r"\((\d+)\s*[Cc]harges?\)", tl))
    if chs:
        return int(chs[-1].group(1))
    lone = list(re.finditer(r"\(\s*(\d+)\s+Charge\)", tl, re.I))
    if lone:
        return int(lone[-1].group(1))
    return None


def extract_bold_bstpet_index(li) -> tuple[str, int | None]:
    b = li.find("b")
    if not b:
        return "", None
    full = re.sub(r"\s+", " ", b.get_text(" ", strip=True))
    mm = re.match(r"^(.+?)\s*\(\s*(\d+)\s*\)\s*$", full.strip())
    if mm:
        return mm.group(1).strip(), int(mm.group(2))
    return full.strip(), None


def clean_ability_name(name: str) -> str:
    name = name.strip().rstrip(":")
    name = re.sub(r"\s*:\s*$", "", name)
    return re.sub(r"\s+", " ", name)


def extract_ability_name(li) -> str:
    b = li.find("b")
    if b:
        if b.get_text(strip=True).lower() == "none":
            return ""
        a = b.find("a")
        raw = (
            a.get_text(strip=True)
            if a and a.get_text(strip=True)
            else re.sub(r"\s*\(\d+\)\s*$", "", b.get_text(" ", strip=True)).strip()
        )
        return clean_ability_name(raw)
    if li.find("b") is None and li.get_text(strip=True).strip().lower() == "none":
        return ""
    a = li.find("a")
    if a and a.get_text(strip=True):
        return clean_ability_name(a.get_text(strip=True))
    return clean_ability_name(li.get_text(" ", strip=True).split(" - ")[0].strip())


def parse_ready_column_from_fragment(html: str, familiar: str) -> tuple[list[str], str]:
    soup = BeautifulSoup(html, "lxml")
    root = soup.find("div", class_=re.compile(r"mw-parser-output")) or soup

    best_ul = None

    for table in root.find_all("table"):
        rows = table.find_all("tr")
        if not rows:
            continue
        header_cells = rows[0].find_all(["th", "td"], recursive=False)
        header_text = " ".join(c.get_text(" ", strip=True) for c in header_cells)
        hl = header_text.lower().replace("\n", " ")
        if "ready" not in hl and "bstpet" not in hl and "abilities" not in hl:
            continue

        best_row_tr = None
        for tr in rows[1:]:
            tds = tr.find_all("td")
            if len(tds) < 2:
                continue
            first = tds[0].get_text(" ", strip=True)
            if not row_matches_familiar(first, familiar):
                continue

            td_ready = None
            for td in reversed(tds):
                if td.find("ul"):
                    td_ready = td
                    break
            if td_ready is None:
                continue

            ul = td_ready.find("ul")
            if ul:
                best_ul = ul
                best_row_tr = tr
                break

        if best_ul:
            break

    if not best_ul:
        return [], "no matching wikitable row with Ready <ul>"

    entries: list[tuple[int, int, str]] = []

    for pos, li in enumerate(best_ul.find_all("li", recursive=False), start=1):
        li_text = li.get_text(" ", strip=True)
        if not li_text:
            continue

        bn = li.find("b")
        if bn and bn.get_text(strip=True).lower() == "none":
            return (["No Ready Abilities"], "wiki table lists None")

        name = extract_ability_name(li)
        tier = extract_charge(li_text)

        abase, bst_idx = extract_bold_bstpet_index(li)
        if abase and "(" in abase:
            abase = re.sub(r"\s*\(.*?\)\s*$", "", abase).strip()

        name = clean_ability_name(name)
        if not name:
            continue

        # Prefer Bold link-derived name slice when Bold had "Name (index)"
        if abase:
            cand = clean_ability_name(re.sub(r"\([^)]*\)\s*$", "", abase.strip()))
            if cand:
                name = cand

        if tier is None:
            tier = 1

        sort_key = bst_idx if bst_idx is not None else pos

        entries.append((sort_key, pos, f"{name} ({tier})"))

    entries.sort(key=lambda x: (x[0], x[1]))
    out = [e[2] for e in entries]

    if not out:
        return [], "wikitable Ready cell parsed zero abilities"

    return out, f"matched; sort={[e[0] for e in entries]}"


def main() -> int:
    sheet = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"e:\Downloads\BST All Jug Pets & Abilities.xlsx")
    delay = float(sys.argv[2]) if len(sys.argv) > 2 else 1.85

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("pip install openpyxl", file=sys.stderr)
        return 1

    wb = load_workbook(sheet)
    ws = wb.active
    hdr = [ws.cell(1, c).value for c in range(1, 25)]

    try:
        col_fam = hdr.index("Familiar") + 1
        col_ready = hdr.index("Ready Abilities (Charges)") + 1
    except ValueError as e:
        print("Unexpected header row", hdr, e, file=sys.stderr)
        return 1

    report_path = sheet.with_name(sheet.stem + "_retail_wiki_sync_report.txt")

    rows_data: list[tuple[int, str]] = []
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, col_fam).value
        if not nm:
            continue
        rows_data.append((r, str(nm).strip()))

    with report_path.open("w", encoding="utf-8") as rep:
        rep.write("row\texcel_name\tresolved_title\tsource_wiki\tstatus\tReady column\tnotes\n")

        for idx, (r, name) in enumerate(rows_data, start=1):
            title = resolve_wiki_title(name)
            frag, src = (None, None)
            if title:
                frag, src = mw_parse_fragment(title)

            if frag is None:
                rep.write(f"{r}\t{name}\t{title or ''}\t\tMISSING_PAGE\t\tno parse on any wiki\n")
                print(f"[{idx}/{len(rows_data)}] {name} SKIP (no page)")
                time.sleep(delay)
                continue

            readies, note = parse_ready_column_from_fragment(frag, name)
            line = (
                ", ".join(re.sub(r"\s{2,}", " ", chunk).strip() for chunk in readies if chunk)
                if readies
                else ""
            )

            if not line.strip():
                rep.write(f'{r}\t{name}\t{title}\t{src}\tPARSE_EMPTY\t\t{note}\n')
                print(f"[{idx}/{len(rows_data)}] {name} EMPTY — {note}")
            else:
                ws.cell(row=r, column=col_ready).value = line
                rep.write(f'{r}\t{name}\t{title}\t{src}\tOK\t{line}\t{note}\n')
                print(f"[{idx}/{len(rows_data)}] {name} -> {line[:110]}")

            time.sleep(delay)

    out = sheet.with_name(sheet.stem + "_RETAIL_WIKi_SYNC.xlsx")
    wb.save(out)
    print("Saved:", out)
    print("Report:", report_path)

    if "--inplace" in sys.argv:
        wb.save(sheet)
        print("Also updated original:", sheet)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
